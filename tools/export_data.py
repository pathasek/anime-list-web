"""
Export Excel anime data to JSON format for web application
"""
import warnings

# openpyxl hlásí u tohoto sešitu neškodná varování o nepodporovaných
# rozšířeních (Slicer, Conditional Formatting) — jen zaplevelují log exportu
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import openpyxl
import json
import os
import sys
import subprocess
import base64
import shutil
import tempfile
import time
from datetime import datetime, date, timedelta
from io import BytesIO
import re
import zipfile
from xml.etree import ElementTree as ET
import win32com.client
import hashlib

import jikan_health


def anime_list_root():
    """Kořen složky `Anime_List` (nadřazená složka celého projektu).

    Odvozuje se z umístění tohoto skriptu:
        Anime_List/Anime List WEB/anime-list-web/tools/export_data.py
        └── 3 úrovně nahoru = Anime_List

    Díky tomu jde celou složku Anime_List přesunout nebo přejmenovat, aniž by
    bylo nutné sahat do skriptů. Pro netypické umístění se dá přebít proměnnou
    prostředí ANIME_LIST_ROOT.
    """
    env = os.environ.get("ANIME_LIST_ROOT")
    if env:
        return env
    tools_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.dirname(os.path.dirname(os.path.dirname(tools_dir)))


def serialize_value(val):
    """Convert Excel values to JSON-serializable format"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, str):
        val_str = val.strip()
        # Match DD.MM.YYYY or D.M.YYYY with optional spaces
        m = re.match(r'^(\d{1,2})\.\s*(\d{1,2})\.\s*(\d{4})$', val_str)
        if m:
            try:
                day = int(m.group(1))
                month = int(m.group(2))
                year = int(m.group(3))
                return datetime(year, month, day).isoformat()
            except ValueError:
                pass
    if isinstance(val, timedelta):
        total_seconds = val.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        return f"{hours}:{minutes:02d}"
    if hasattr(val, 'value'):  # ArrayFormula or similar
        return str(val)
    return val

def extract_hyperlink_text(cell):
    """Extract display text from hyperlink if present"""
    if cell.hyperlink:
        return cell.value if cell.value else str(cell.hyperlink.target)
    return cell.value

def export_anime_list(wb, wb_comments=None):
    """Export main anime list"""
    ws = wb["ANIME LIST"]
    data = []
    
    # Load tags from cache
    tags_cache = {}
    if "MAL Cache + Interactive Rating" in wb.sheetnames:
        try:
            ws_cache = wb["MAL Cache + Interactive Rating"]
            
            # 1. Load Tag Descriptions (BP=68, BQ=69)
            tag_descriptions = {}
            for row in range(2, ws_cache.max_row + 1):
                t_name = ws_cache.cell(row, 68).value
                t_desc = ws_cache.cell(row, 69).value
                if t_name and t_desc:
                    tag_descriptions[str(t_name).strip()] = str(t_desc).strip()
                    
            # 2. Map tags to Anime (BY=77, BZ=78)
            for row in range(2, ws_cache.max_row + 1):
                name_val = ws_cache.cell(row, 77).value
                tags_val = ws_cache.cell(row, 78).value
                if name_val and tags_val:
                    name_str = str(name_val).strip()
                    new_tags = []
                    for t in str(tags_val).strip().split(';'):
                        parts = t.split(':')
                        if len(parts) >= 1:
                            t_n = parts[0].strip()
                            rank = parts[1].strip() if len(parts) > 1 else "0"
                            t_d = tag_descriptions.get(t_n, "").replace(':', '-')
                            new_tags.append(f"{t_n}:{rank}:{t_d}")
                    tags_cache[name_str] = ";".join(new_tags)
        except Exception as e:
            print(f"  Warning: Error reading tags from cache: {e}")
    
    # Headers are in row 1
    headers = [
        "index", "thumbnail", "name", "type", "studio", "release_date",
        "themes", "genres", "episodes", "episode_duration", "rating",
        "start_date", "end_date", "rewatch_count", "total_time",
        "dub", "status", "mal_url"
    ]
    
    for row in range(2, ws.max_row + 1):
        # Skip empty rows
        if not ws.cell(row, 3).value:  # Name column
            continue
            
        anime = {}
        col_map = {
            1: "index",
            2: "thumbnail",
            3: "name",
            4: "type",
            5: "studio",
            6: "release_date",
            7: "themes",
            8: "genres",
            9: "episodes",
            10: "episode_duration",
            11: "rating",
            12: "start_date",
            13: "end_date",
            14: "rewatch_count",
            15: "total_time",
            16: "dub",
            17: "status"
        }
        
        for col, key in col_map.items():
            cell = ws.cell(row, col)
            val = cell.value
            
            # Special user rule for Status (Column 17/Q)
            if col == 17:
                if not val or str(val).strip() == "":
                    anime[key] = "PENDING"
                elif str(val).strip().upper() in ["AIRING", "AIRING!"]:
                    anime[key] = "AIRING!"
                else:
                    anime[key] = "FINISHED"
                continue

            # Handle hyperlinks in studio column
            if col == 5 and cell.hyperlink:
                val = str(cell.value) if cell.value else ""
                # Extract studio name from hyperlink text
                if val.startswith("=HYPERLINK"):
                    parts = val.split(",")
                    if len(parts) > 1:
                        val = parts[1].strip().strip('"').strip(")")
            
            anime[key] = serialize_value(val)
        
        # Extract MAL URL from name column hyperlink
        name_cell = ws.cell(row, 3)
        if name_cell.hyperlink and name_cell.hyperlink.target:
            anime["mal_url"] = name_cell.hyperlink.target
        else:
            anime["mal_url"] = None
        
        if anime.get("studio") and isinstance(anime["studio"], str):
            if anime["studio"].startswith("=HYPERLINK"):
                # Try to extract from the formula
                studio = ws.cell(row, 5).value
                if hasattr(studio, '__str__'):
                    anime["studio"] = str(studio).split(",")[-1].strip().strip('"').strip(")")
        
        # Extract series and rewatches from comment if possible
        series_name = None
        rewatch_list = []
        if wb_comments:
            try:
                ws_comments = wb_comments["ANIME LIST"]
                comment = ws_comments.cell(row, 3).comment
                if comment:
                    text = comment.text
                    for line in text.split('\n'):
                        line = line.strip()
                        if not line: continue
                        
                        if 'Rewatch' in line:
                            rewatch_list.append(line)
                            
                        # Use split and case-insensitive check for "Název série"
                        if ":" in line:
                            parts = line.split(":", 1)
                            if parts[0].strip().lower() == "název série":
                                series_name = parts[1].strip()
            except Exception as e:
                print(f"  Warning: Error extracting comment data at row {row}: {e}")
        
        anime["series"] = series_name
        anime["rewatches"] = rewatch_list
        
        anime_name_str = str(anime.get("name", "")).strip()
        anime["tags"] = tags_cache.get(anime_name_str)
        # Fallback: pokud se tagy nenašly podle jména (např. uživatel přejmenoval
        # "Witch Hat Atelier" → "Witch Hat Atelier, S01"), zkusíme název série
        if not anime["tags"] and anime.get("series"):
            series_str = str(anime["series"]).strip()
            anime["tags"] = tags_cache.get(series_str)
        
        data.append(anime)
    
    return data

def export_history_log(wb, wb_comments=None):
    """Export watching history"""
    ws = wb["HISTORY LOG"]
    data = []
    
    # Map comments for column A (Anime column) if wb_comments is provided
    rewatch_map = {}
    if wb_comments:
        ws_comments = wb_comments["HISTORY LOG"]
        for row in range(3, ws_comments.max_row + 1):
            cell = ws_comments.cell(row=row, column=1)
            if cell.comment:
                comment_text = cell.comment.text
                # Look for "X. Rewatch" format
                match = re.search(r'(\d+)\.\s*Rewatch', comment_text, re.IGNORECASE)
                if match:
                    rewatch_map[row] = match.group(1)

    current_date = None
    
    for row in range(3, ws.max_row + 1):  # Skip headers
        name = ws.cell(row, 1).value
        episodes = ws.cell(row, 2).value
        time_spent = ws.cell(row, 3).value
        date_val = ws.cell(row, 4).value
        
        if date_val:
            current_date = serialize_value(date_val)
        
        # Skip if no name or episodes
        if not name or not episodes:
            continue
            
        # Skip summary rows - these start with "(" like "(2x)" which indicates daily count
        name_str = str(name).strip()
        if name_str.startswith("(") and name_str.endswith(")"):
            continue
        
        # Also skip if episodes column is just a count like "(14x)" without episode details
        eps_str = str(episodes).strip()
        if eps_str.startswith("(") and eps_str.endswith(")") and "EP" not in eps_str.upper():
            continue
            
        entry = {
            "name": serialize_value(name),
            "episodes": serialize_value(episodes),
            "time": serialize_value(time_spent),
            "date": current_date,
            "rewatch": rewatch_map.get(row)
        }
        data.append(entry)
    
    return data

def export_general_stats(wb, wb_comments=None):
    """Export general statistics from OBECNÉ INFORMACE"""
    ws = wb["OBECNÉ INFORMACE"]
    
    # Dynamic year detection: scan row 2 from column G onward for year headers
    years = ["total"]  # Column F is always total
    col = 7  # Start at G
    while True:
        val = ws.cell(2, col).value
        if val is None:
            break
        year_str = str(int(val)) if isinstance(val, (int, float)) else str(val).strip()
        if year_str.isdigit() and len(year_str) == 4:
            years.append(year_str)
        else:
            break
        col += 1
    
    # If no years detected, fall back to known years
    if len(years) == 1:
        years = ["total", "2024", "2025", "2026"]
    
    # Extract EXACT dashboard table parity (D2:I9)
    # D is index 4, E is index 5 (skip), F is index 6, G=7, H=8, I=9
    # Limit rows to 9 to skip the Type breakdown arrays as requested by user
    dashboard_table = []
    
    # Read headers from row 2 (D2, F2, G2, H2, I2...)
    headers = [serialize_value(ws.cell(2, 4).value)]
    for c in range(6, 6 + len(years)):
        val = ws.cell(2, c).value
        headers.append(serialize_value(val))
    dashboard_table.append(headers)
    
    # Read data from rows 3 to 9
    for r in range(3, 10):
        row_data = [serialize_value(ws.cell(r, 4).value)] # Row label
        
        for c in range(6, 6 + len(years)):
            cell_val = ws.cell(r, c).value
            
            # Row 3 is "Čas sledování (hh:mm)" which is stored in days in Excel
            if r == 3 and isinstance(cell_val, (int, float)):
                total_hours = float(cell_val) * 24
                hrs = int(total_hours)
                mins = int((total_hours - hrs) * 60)
                cell_text = f"{hrs}:{mins:02d}"
            # Formatting numeric values
            elif isinstance(cell_val, (int, float)):
                if int(cell_val) != cell_val:
                    # Format as float with 2 decimal places, replacing . with ,
                    cell_text = f"{cell_val:.2f}".replace('.', ',')
                    if cell_text.endswith(',00'):
                        cell_text = str(int(cell_val))
                else:
                    cell_text = str(int(cell_val))
            else:
                cell_text = serialize_value(cell_val) or "-"
                
            row_data.append(cell_text)
            
        dashboard_table.append(row_data)

    stats = {
        "last_update": serialize_value(ws.cell(2, 1).value),
        "dashboard_table": dashboard_table,
        "total_time": {},
        "total_episodes": {},
        "avg_episode_duration": {},
        "anime_count": {},
        "comments": {
            "total_time": {},
            "total_episodes": {},
            "rewatch_count": {}
        }
    }
    
    for i, year in enumerate(years):
        col = 6 + i  # Columns F, G, H, I...
        stats["total_time"][year] = serialize_value(ws.cell(3, col).value)
        stats["total_episodes"][year] = serialize_value(ws.cell(5, col).value)
        stats["avg_episode_duration"][year] = serialize_value(ws.cell(6, col).value)
    
    # Extract comments from the formula workbook (if available)
    if wb_comments:
        try:
            ws_c = wb_comments["OBECNÉ INFORMACE"]
            for i, year in enumerate(years):
                col = 6 + i
                # Row 3: total time comments (rewatch breakdown)
                cell_time = ws_c.cell(3, col)
                if cell_time.comment:
                    stats["comments"]["total_time"][year] = cell_time.comment.text
                # Row 5: total episodes comments (rewatch breakdown)
                cell_eps = ws_c.cell(5, col)
                if cell_eps.comment:
                    stats["comments"]["total_episodes"][year] = cell_eps.comment.text
                # Row 7: rewatch count comments (rewatch list)
                cell_rw = ws_c.cell(7, col)
                if cell_rw.comment:
                    stats["comments"]["rewatch_count"][year] = cell_rw.comment.text
        except Exception as e:
            print(f"  Warning: Could not extract comments: {e}")
    
    # Remove empty comment sections
    stats["comments"] = {k: v for k, v in stats["comments"].items() if v}
    if not stats["comments"]:
        del stats["comments"]
    
    return stats

def export_favorites(wb):
    """Export favorite OP/ED/OST"""
    ws = wb["ANIME FAV OP + ED + OST"]
    data = []
    
    # Actual Excel column layout (verified from headers in row 2):
    # Col 8  = Pořadí (index)
    # Col 9  = Název Anime
    # Col 10 = Typ (OP/ED/OST)
    # Col 11 = Song
    # Col 12 = Autor
    # Col 13 = Jazyk
    # Col 14 = Hodnocení textu (Lyrics)
    # Col 15 = Emoce (Emotion)
    # Col 16 = Melodie (Melody)
    # Col 17 = Videoklip (Video)
    # Col 18 = Kvalita hlasu (Voice quality)
    # Col 19 = Sing-along faktor
    # Col 20 = Frisson feeling (Ano/Ne)
    # Col 21 = Hodnocení (průměrné) — average of sub-ratings
    # Col 22 = Finální hodnocení — final user rating
    
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row, 9).value  # Column I - Název Anime
        if not name:
            continue
            
        # Parse frisson as boolean
        frisson_val = ws.cell(row, 20).value
        has_frisson = str(frisson_val).strip().lower() in ('ano', 'yes', 'true') if frisson_val else False
            
        entry = {
            "index": serialize_value(ws.cell(row, 8).value),
            "anime_name": serialize_value(name),
            "type": serialize_value(ws.cell(row, 10).value),
            "song": serialize_value(ws.cell(row, 11).value),
            "author": serialize_value(ws.cell(row, 12).value),
            "language": serialize_value(ws.cell(row, 13).value),
            "rating_lyrics": serialize_value(ws.cell(row, 14).value),
            "rating_emotion": serialize_value(ws.cell(row, 15).value),
            "rating_melody": serialize_value(ws.cell(row, 16).value),
            "rating_video": serialize_value(ws.cell(row, 17).value),
            "rating_voice": serialize_value(ws.cell(row, 18).value),
            "sing_along": serialize_value(ws.cell(row, 19).value),
            "has_frisson": has_frisson,
            "rating_avg": serialize_value(ws.cell(row, 21).value),
            "rating_final": serialize_value(ws.cell(row, 22).value),
        }
        data.append(entry)
    
    return data

def export_ost_tables(wb, wb_formulas=None):
    """Export the three OST tables: scenes, pieces, whole"""
    ws = wb["ANIME FAV OP + ED + OST"]
    ws_formulas = wb_formulas["ANIME FAV OP + ED + OST"] if wb_formulas else None
    
    def get_val_and_link(r, c):
        cell_data = ws.cell(row=r, column=c)
        cell_formula = ws_formulas.cell(row=r, column=c) if ws_formulas else None
        
        # Get link
        link = None
        if hasattr(cell_data, 'hyperlink') and cell_data.hyperlink and cell_data.hyperlink.target:
            link = cell_data.hyperlink.target
        elif cell_formula and hasattr(cell_formula, 'hyperlink') and cell_formula.hyperlink and cell_formula.hyperlink.target:
            link = cell_formula.hyperlink.target
            
        formula_str = str(cell_formula.value) if cell_formula and cell_formula.value else ""
        data_str = str(cell_data.value) if cell_data and cell_data.value else ""
        
        if not link and formula_str.upper().startswith("=HYPERLINK"):
            import re
            m = re.search(r'=HYPERLINK\("([^"]+)"', formula_str, re.IGNORECASE)
            if m: link = m.group(1)
        if not link and data_str.upper().startswith("=HYPERLINK"):
            import re
            m = re.search(r'=HYPERLINK\("([^"]+)"', data_str, re.IGNORECASE)
            if m: link = m.group(1)
            
        # Get text
        val = cell_data.value
        if isinstance(val, str) and val.upper().startswith("=HYPERLINK"):
            import re
            m = re.search(r'=HYPERLINK\("[^"]+",\s*"([^"]+)"\)', val, re.IGNORECASE)
            if m: 
                val = m.group(1)
            else:
                m2 = re.search(r'=HYPERLINK\("[^"]+",\s*(.+)\)', val, re.IGNORECASE)
                if m2: 
                    val = str(m2.group(1)).strip().strip('"')
                
        if val is None and cell_formula and isinstance(cell_formula.value, str) and not cell_formula.value.startswith("="):
            val = cell_formula.value
            
        return serialize_value(val), link

    data = {
        "scenes": [],
        "pieces": [],
        "whole": []
    }
    
    # Table_FAV_OST_SCENES: Y(25), Z(26), AA(27)
    for row in range(3, ws.max_row + 1):
        anime, anime_link = get_val_and_link(row, 25)
        if not anime: continue
        episode, ep_link = get_val_and_link(row, 26)
        scene, scene_link = get_val_and_link(row, 27)
        data["scenes"].append({
            "anime_name": anime, "anime_url": anime_link,
            "episode": episode, "episode_url": ep_link,
            "scene": scene, "scene_url": scene_link
        })
        
    # Table_FAV_OST_PIECES: AD(30), AE(31)
    for row in range(3, ws.max_row + 1):
        anime, anime_link = get_val_and_link(row, 30)
        if not anime: continue
        ost, ost_link = get_val_and_link(row, 31)
        data["pieces"].append({
            "anime_name": anime, "anime_url": anime_link,
            "ost_name": ost, "ost_url": ost_link
        })
        
    # Table_FAV_OST_WHOLE: AG(33)=Pořadí (order), AH(34)=anime, AI(35)=YT, AJ(36)=Spotify
    # POZOR: pole "order" je NUTNÉ — frontend (Favorites.jsx) podle něj řadí "OST Only (As a Whole)";
    # bez něj spadne na abecední řazení (localeCompare). AG obsahuje '01.', '02.', ...
    for row in range(3, ws.max_row + 1):
        order_val, _ = get_val_and_link(row, 33)
        anime, anime_link = get_val_and_link(row, 34)
        if not anime: continue
        yt, yt_link = get_val_and_link(row, 35)
        spotify, spotify_link = get_val_and_link(row, 36)
        data["whole"].append({
            "order": order_val,
            "anime_name": anime, "anime_url": anime_link,
            "yt_playlist": yt, "yt_url": yt_link,
            "spotify_playlist": spotify, "spotify_url": spotify_link
        })
        
    return data

def export_category_ratings(wb):
    """Export category ratings for each anime from MAL Cache sheet"""
    ws = wb["MAL Cache + Interactive Rating"]
    
    # Data structure: anime_name -> {category: rating}
    ratings = {}
    
    # Categories are in columns A-D: Název Anime, Typ, Položka, Hodnocení
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        typ = ws.cell(row, 2).value
        category = ws.cell(row, 3).value
        rating = ws.cell(row, 4).value
        
        if not name or not category or not rating:
            continue
            
        # Only process category ratings (not episodes)
        if str(typ).strip() != "Kategorie":
            continue
            
        # Clean category name (remove weird prefixes)
        cat_str = str(category).strip()
        if cat_str.startswith("_x000D_"):
            cat_str = cat_str.replace("_x000D_", "").strip()
        if cat_str.startswith("\n"):
            cat_str = cat_str[1:].strip()
            
        name_str = str(name).strip()
        
        if name_str not in ratings:
            ratings[name_str] = {}
        
        try:
            ratings[name_str][cat_str] = float(rating) if isinstance(rating, (int, float)) else float(str(rating).replace(",", "."))
        except:
            pass
    
    # Convert to list format
    data = []
    for anime_name, categories in ratings.items():
        if categories:  # Only include anime with ratings
            data.append({
                "name": anime_name,
                "categories": categories
            })
    
    return data

def export_episode_ratings(wb):
    """Export episode ratings for each anime from MAL Cache sheet"""
    ws = wb["MAL Cache + Interactive Rating"]
    
    # Data structure: anime_name -> [{episode: "EP 1", rating: 7.5}, ...]
    ratings = {}
    
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        typ = ws.cell(row, 2).value
        episode = ws.cell(row, 3).value
        rating = ws.cell(row, 4).value
        
        if not name or not episode:
            continue
            
        # Only process episode ratings
        if str(typ).strip() != "Epizoda":
            continue
            
        name_str = str(name).strip()
        ep_str = str(episode).strip()
        
        if name_str not in ratings:
            ratings[name_str] = []
        
        try:
            rating_val = float(rating) if isinstance(rating, (int, float)) else float(str(rating).replace(",", "."))
            ratings[name_str].append({
                "episode": ep_str,
                "rating": rating_val
            })
        except:
            pass
    
    # Convert to list format and sort episodes
    data = []
    for anime_name, episodes in ratings.items():
        if episodes:
            # Sort by episode number
            sorted_eps = sorted(episodes, key=lambda x: int(x["episode"].replace("EP ", "").strip()) if x["episode"].replace("EP ", "").strip().isdigit() else 0)
            data.append({
                "name": anime_name,
                "episodes": sorted_eps
            })

    data = _zahod_zkopirovane_bloky(wb, data)

    return data


def _zahod_zkopirovane_bloky(wb, data):
    """Zahodí známky epizod, které jsou kopií jiného titulu.

    V listu „MAL Cache + Interactive Rating" se stane, že se řádky jednoho
    anime omylem zkopírují pod název jiného. Poznávací znamení: dva různé
    tituly mají naprosto shodnou posloupnost známek. Rozhodčím je počet epizod
    v hlavním seznamu: blok, který mu odpovídá, zůstane, ten druhý padá.

    Reálný případ, kvůli kterému to vzniklo: film „The Disappearance of Haruhi
    Suzumiya" (1 epizoda) nesl 14 známek zkopírovaných ze seriálu
    „The Melancholy of Haruhi Suzumiya, S02". Web z toho počítal průměr epizod
    a analytickou tenzi z cizích dat.

    Schválně se nekontroluje jen počet epizod proti seznamu. U právě
    vysílaných sérií je běžné, že mám ohodnoceno víc dílů, než kolik seznam
    eviduje, a to není chyba. Zahazuje se výhradně shodná posloupnost.
    """
    pocty_epizod = {}
    try:
        ws = wb["ANIME LIST"]
        for row in range(2, ws.max_row + 1):
            nazev = ws.cell(row, 3).value        # sloupec C, název
            pocet = ws.cell(row, 9).value        # sloupec I, počet epizod
            if nazev and isinstance(pocet, (int, float)):
                pocty_epizod[str(nazev).strip()] = int(pocet)
    except Exception as e:
        print(f"  [kontrola epizod] hlavni seznam nelze precist, kontrola se preskakuje: {e}")
        return data

    podle_otisku = {}
    for zaznam in data:
        otisk = tuple(e["rating"] for e in zaznam["episodes"])
        if len(otisk) < 3:
            continue
        podle_otisku.setdefault(otisk, []).append(zaznam)

    k_zahozeni = set()
    for skupina in podle_otisku.values():
        if len(skupina) < 2:
            continue
        for zaznam in skupina:
            ocekavano = pocty_epizod.get(zaznam["name"])
            if ocekavano is not None and len(zaznam["episodes"]) != ocekavano:
                k_zahozeni.add(zaznam["name"])
                print(f"  [kontrola epizod] VAROVANI: '{zaznam['name']}' ma "
                      f"{len(zaznam['episodes'])} znamek, ale v seznamu ma "
                      f"{ocekavano} epizod, a posloupnost je shodna s jinym "
                      f"titulem. Blok se do exportu nedostane. Oprav to "
                      f"v Excelu, tohle je jen zachytna sit.")

    if k_zahozeni:
        data = [z for z in data if z["name"] not in k_zahozeni]

    return data

def export_notes(wb):
    """Export narrative reviews/notes for each anime from MAL Cache sheet"""
    ws = wb["MAL Cache + Interactive Rating"]
    
    # Data structure: list of {name, note}
    data = []
    
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        typ = ws.cell(row, 2).value
        note = ws.cell(row, 3).value
        
        if not name or not note:
            continue
            
        # Only process notes/reviews
        if str(typ).strip() != "Poznámka":
            continue
            
        name_str = str(name).strip()
        note_str = str(note).strip()
        
        if note_str:
            data.append({
                "name": name_str,
                "note": note_str
            })
    
    return data

def export_plan_to_watch(wb):
    """Export Plan to Watch list"""
    ws = wb["ANIME PLAN TO WATCH + FUTURES"]
    data = []
    
    # Dynamically detect last row (like VBA: ws.Cells(ws.Rows.Count, "G").End(xlUp).Row)
    last_row = ws.max_row
    
    # Plan to Watch table starts at column G, row 3
    for row in range(3, last_row + 1):
        name = ws.cell(row, 7).value  # Column G - Název
        if not name:
            continue
        
        # Filter: must have "Pořadí" in column F to distinguish from category headers (like VBA)
        poradi = ws.cell(row, 6).value  # Column F
        if poradi is None or (isinstance(poradi, str) and poradi.strip() == ""):
            continue
            
        # Read total time from column K (in minutes, like VBA COL_CAS_TOTAL)
        total_time_val = ws.cell(row, 11).value  # Column K
        total_time = None
        if total_time_val is not None and not isinstance(total_time_val, str):
            try:
                total_time = float(total_time_val)
            except (ValueError, TypeError):
                pass
        
        entry = {
            "name": serialize_value(name),
            "type": serialize_value(ws.cell(row, 8).value),   # Column H - Typ
            "episodes": serialize_value(ws.cell(row, 9).value),  # Column I - Počet epizod
            "total_time": total_time,                            # Column K - Celkový čas (minuty)
            "source": serialize_value(ws.cell(row, 12).value),   # Column L - Důvod/Zdroj
            "notes": serialize_value(ws.cell(row, 13).value)     # Column M - Status (Vydáno/AIRING!)
        }
        data.append(entry)
    
    return data

def get_file_hash(filepath):
    """Return MD5 hash of a file to detect changes."""
    hasher = hashlib.md5()
    try:
        with open(filepath, 'rb') as f:
            buf = f.read(65536)
            while len(buf) > 0:
                hasher.update(buf)
                buf = f.read(65536)
        return hasher.hexdigest()
    except:
        return None

def _parsuj_alt_text(alt_text):
    """Rozparsuje alt-text tvaru na CHAR_ID / ANIME_NAME / NAME.

    Stejná logika jako dřívější čtení přes Excel COM, jen vytažená stranou,
    aby ji mohla používat i cesta přes zip i COM fallback."""
    parsed_data = {}
    if "CHAR_ID:" in alt_text or "ANIME_NAME:" in alt_text:
        char_match = re.search(r'CHAR_ID:(\d+)', alt_text)
        if char_match:
            parsed_data["CHAR_ID"] = char_match.group(1)

        anime_match = re.search(r'ANIME_NAME:(.+?)(?=;[A-Z_]+:|$)', alt_text)
        if anime_match:
            parsed_data["ANIME_NAME"] = anime_match.group(1).strip()

        name_match = re.search(r'NAME:(.+?)(?=;[A-Z_]+:|$)', alt_text)
        if name_match:
            parsed_data["NAME"] = name_match.group(1).strip()
    else:
        parsed_data["NAME"] = alt_text.strip()
    return parsed_data


def _najdi_drawing_xml(zf):
    """Vrátí cestu k drawing XML listu OBECNÉ INFORMACE uvnitř xlsm zipu."""
    xml_content = zf.read('xl/workbook.xml')
    match = re.search(rb'<sheet[^>]+name="OBEC[^"]+"[^>]+r:id="([^"]+)"', xml_content)
    if not match:
        raise Exception("Sheet OBECNE INFORMACE not found in XML")
    sheet_rid = match.group(1).decode('utf-8')

    xml_content = zf.read('xl/_rels/workbook.xml.rels')
    match = re.search(rb'<Relationship[^>]+Id="' + sheet_rid.encode() + rb'"[^>]+Target="([^"]+)"', xml_content)
    sheet_path = 'xl/' + match.group(1).decode('utf-8')

    xml_content = zf.read(sheet_path)
    match = re.search(rb'<drawing r:id="([^"]+)"', xml_content)
    if not match:
        raise Exception("No drawing found in sheet")
    drawing_rid = match.group(1).decode('utf-8')

    rels_path = sheet_path.replace('worksheets/', 'worksheets/_rels/') + '.rels'
    xml_content = zf.read(rels_path)
    match = re.search(rb'<Relationship[^>]+Id="' + drawing_rid.encode() + rb'"[^>]+Target="([^"]+)"', xml_content)
    return 'xl/' + match.group(1).decode('utf-8').replace('../', '')


def _shapes_pres_com_fallback(wb_path):
    """Záložní čtení alt-textů přes Excel COM (původní cesta).

    Použije se jen když zip nevydá žádné tvary a neběžíme orchestrovaně —
    při souběhu s Excel makrem by se COM mohl chytit cizí instance a viset."""
    print("  COM fallback: Connecting to Excel to read shape Alternative Text...")
    shapes_data = {}
    excel_was_open = False
    xl = None
    try:
        try:
            xl = win32com.client.GetActiveObject("Excel.Application")
            excel_was_open = True
        except:
            xl = win32com.client.Dispatch('Excel.Application')

        wb_com = None
        for w in xl.Workbooks:
            if w.FullName.lower() == wb_path.lower():
                wb_com = w
                excel_was_open = True
                break

        if not wb_com:
            wb_com = xl.Workbooks.Open(wb_path, ReadOnly=True)

        ws_com = wb_com.Sheets('OBECNÉ INFORMACE')

        def get_all_shapes(shapes_collection):
            found_shapes = []
            for shape in shapes_collection:
                try:
                    name = str(shape.Name)
                    if name.startswith('Top10_') or name.startswith('HM_'):
                        found_shapes.append(shape)
                    if shape.Type == 6:  # msoGroup
                        found_shapes.extend(get_all_shapes(shape.GroupItems))
                except:
                    pass
            return found_shapes

        for shape in get_all_shapes(ws_com.Shapes):
            name = str(shape.Name)
            if name.startswith('HM_Char_'):
                continue
            alt_text = shape.AlternativeText
            if alt_text:
                shapes_data[name] = {
                    "shape_name": name,
                    "data": _parsuj_alt_text(alt_text),
                    "image_file": None
                }

        if not excel_was_open:
            wb_com.Close(SaveChanges=False)
            xl.Quit()

        print(f"  Found {len(shapes_data)} valid shapes via COM.")
    except Exception as e:
        print(f"  Error reading shapes via COM: {e}")
        try:
            if not excel_was_open and xl is not None:
                xl.Quit()
        except:
            pass
    return shapes_data


def export_top_favorites(wb_path, output_dir):
    """
    Export Top 10 and HM Anime/Characters data. Alt-texty tvarů i vložené
    obrázky se čtou přímo ze zipu sešitu (bez spouštění Excelu), COM zůstává
    jen jako fallback.
    """
    images_dir = os.path.join(output_dir, "..", "images", "top_favorites")
    os.makedirs(images_dir, exist_ok=True)
    json_path = os.path.join(output_dir, "top_favorites.json")
    hash_path = os.path.join(output_dir, "top_favorites_hash.txt")
    
    # Check if we need to update
    current_hash = get_file_hash(wb_path)
    old_hash = None
    if os.path.exists(hash_path):
        try:
            with open(hash_path, 'r') as f:
                old_hash = f.read().strip()
        except:
            pass
            
    # If the file hasn't changed and the json exists, skip extraction
    if current_hash == old_hash and os.path.exists(json_path) and os.listdir(images_dir):
        print("  Excel file hasn't changed. Skipping heavy Top Favorites extraction.")
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            print("  Failed to load cached JSON, re-extracting...")
            pass

    # Alt-texty tvarů se čtou přímo ze zipu sešitu (atribut descr elementu
    # cNvPr v drawing XML). Ověřeno proti Excel COM AlternativeText: na všech
    # 73 tvarech znak po znaku shodné, včetně parsování CHAR_ID/ANIME_NAME.
    # Excel se tak vůbec nespouští a export smí běžet souběžně s makrem.
    print("  Reading shape alt texts + images from workbook zip (bez Excelu)...")

    shapes_data = {}
    shape_embeds = {}

    # Copy file to temp just to be safe while unzipping
    temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsm")
    os.close(temp_fd)

    try:
        shutil.copy2(wb_path, temp_path)

        with zipfile.ZipFile(temp_path, 'r') as zf:
            drawing_path = _najdi_drawing_xml(zf)
            root = ET.fromstring(zf.read(drawing_path))

            R_NS = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
            for el in root.iter():
                tag = el.tag.rsplit('}', 1)[-1]
                if tag not in ('sp', 'pic'):
                    continue
                cnv = None
                embed = None
                for sub in el.iter():
                    st = sub.tag.rsplit('}', 1)[-1]
                    if st == 'cNvPr' and cnv is None:
                        cnv = sub
                    elif st == 'blip' and embed is None:
                        embed = sub.get(R_NS + 'embed')
                if cnv is None:
                    continue
                name = cnv.get('name') or ''
                if not (name.startswith('Top10_') or name.startswith('HM_')):
                    continue
                # Exclude HM Characters as requested by user
                if name.startswith('HM_Char_'):
                    continue
                if embed and not name.startswith(('Top10_Anime_', 'Top10_CharAnime_')):
                    # Vložené obrázky Top10 anime se nevytahují: image_file jim
                    # vzápětí přepíše poster z Jikanu (Jikan_*.jpg). Dřívější
                    # regexové čtení je kvůli seskupení tvarů stejně nevytáhlo,
                    # takže by šlo jen o mrtvé soubory navíc v repozitáři.
                    shape_embeds[name] = embed
                alt_text = cnv.get('descr')
                if alt_text:
                    shapes_data[name] = {
                        "shape_name": name,
                        "data": _parsuj_alt_text(alt_text),
                        "image_file": None  # Will be mapped below
                    }

            if not shapes_data:
                print("  Warning: zip nevydal žádné alt-texty tvarů.")
                if os.environ.get("NLM_ORCHESTRATED"):
                    # Souběžný běh s Excel makrem: COM fallback je zakázaný,
                    # ať se export nechytí cizí instance Excelu. Volající
                    # zachová minulý top_favorites.json (viz has_items check).
                    print("  Orchestrovaný běh: COM fallback zakázán, končím bez dat.")
                    return {"top10_anime": [], "hm_anime": [], "top10_chars": []}
                shapes_data = _shapes_pres_com_fallback(wb_path)
                if not shapes_data:
                    return {"top10_anime": [], "hm_anime": [], "top10_chars": []}
            else:
                print(f"  Found {len(shapes_data)} valid shapes (zip descr).")

            # Extrakce originálních obrázků (stejné rels mapování jako dřív)
            drw_rels_path = drawing_path.replace('drawings/', 'drawings/_rels/') + '.rels'
            xml_content = zf.read(drw_rels_path).decode('utf-8')

            rid_map = {}
            rels = re.finditer(r'<Relationship[^>]+Id="([^"]+)"[^>]+Target="([^"]+)"', xml_content)
            for rel in rels:
                rid_map[rel.group(1)] = rel.group(2).replace('../', '')

            for name, embed_rid in shape_embeds.items():
                if name not in shapes_data:
                    continue
                if embed_rid in rid_map:
                    image_zip_path = 'xl/' + rid_map[embed_rid]
                    ext = image_zip_path.split('.')[-1]
                    output_file_name = f"{name}.{ext}"
                    output_file_path = os.path.join(images_dir, output_file_name)

                    try:
                        with zf.open(image_zip_path) as img_file:
                            img_bytes = img_file.read()
                        # Nezměněný obrázek se nepřepisuje: zachová se mtime,
                        # takže build_top_favorites_thumbs nepřegenerovává
                        # zmenšeniny, které se reálně nezměnily.
                        stejny = False
                        if os.path.exists(output_file_path):
                            try:
                                with open(output_file_path, 'rb') as f_old:
                                    stejny = f_old.read() == img_bytes
                            except Exception:
                                stejny = False
                        if not stejny:
                            with open(output_file_path, 'wb') as out_file:
                                out_file.write(img_bytes)
                        shapes_data[name]["image_file"] = f"images/top_favorites/{output_file_name}"
                    except Exception as e:
                        print(f"  Warning: failed to extract {image_zip_path}: {e}")

    except Exception as e:
        print(f"  Error extracting shapes/images via Zip: {e}")
        if not shapes_data:
            return {"top10_anime": [], "hm_anime": [], "top10_chars": []}
    finally:
        os.remove(temp_path)

    # Categorize into lists and sort them by rank
    # Rank is the number at the end, e.g. "Top10_Char_5" -> 5
    def get_rank(name):
        try:
            return int(name.split('_')[-1])
        except:
            return 999
            
    top10_anime = []
    hm_anime = []
    top10_chars = []
    
    for name, sdata in shapes_data.items():
        if name.startswith('Top10_Anime_') or name.startswith('Top10_CharAnime_'):
            top10_anime.append(sdata)
        elif name.startswith('HM_Anime_'):
            hm_anime.append(sdata)
        elif name.startswith('Top10_Char_'):
            top10_chars.append(sdata)
            
    top10_anime.sort(key=lambda x: get_rank(x["shape_name"]))
    hm_anime.sort(key=lambda x: get_rank(x["shape_name"]))
    top10_chars.sort(key=lambda x: get_rank(x["shape_name"]))

    print("  Loading anime_list.json for accurate MAL ID lookups...")
    anime_list_data = []
    try:
        anime_list_path = os.path.join(output_dir, "anime_list.json")
        with open(anime_list_path, 'r', encoding='utf-8') as f:
            anime_list_data = json.load(f)
    except Exception as e:
        print(f"  Warning: Could not load anime_list.json for ID matching: {e}")
        pass
        
    def get_mal_id(anime_name):
        # find in anime_list_data
        search = anime_name.lower().strip().replace('  ', ' ')
        for a in anime_list_data:
            a_name = (a.get('name') or '').lower().strip().replace('  ', ' ')
            a_series = (a.get('series') or '').lower().strip().replace('  ', ' ')
            
            # Match name exactly OR match series exactly
            if a_name == search or a_series == search:
                url = a.get('mal_url', '')
                if url:
                    import re
                    match = re.search(r'/anime/(\d+)', url)
                    if match:
                        return match.group(1)
                        
        # Secondary fallback matches: starts with
        for a in anime_list_data:
            a_name = (a.get('name') or '').lower().strip()
            if a_name.startswith(search):
                url = a.get('mal_url', '')
                if url:
                    import re
                    match = re.search(r'/anime/(\d+)', url)
                    if match:
                        return match.group(1)
        return None

    print("  Fetching Top 10 Anime imagery from Jikan API by MAL ID (Skipping HM Anime)...")
    import requests

    for sdata in top10_anime:
        anime_name = sdata["data"].get("NAME") or sdata["data"].get("ANIME_NAME")
        if anime_name:
            safe_name = "".join([c for c in anime_name if c.isalpha() or c.isdigit() or c in (' ', '-', '_')]).strip()
            if not safe_name:
                safe_name = "unknown"
                
            output_file_name = f"Jikan_{safe_name}.jpg"
            output_file_path = os.path.join(images_dir, output_file_name)
            
            sdata["image_file"] = f"images/top_favorites/{output_file_name}"
            
            if not os.path.exists(output_file_path):
                if jikan_health.je_vypadek():
                    print(f"    Jikan je mimo provoz (sdílená pojistka), poster pro {anime_name} se zkusí příště.")
                    continue
                print(f"    Fetching Jikan API for: {anime_name}")
                mal_id = get_mal_id(anime_name)
                
                try:
                    if mal_id:
                        print(f"      Matched MAL ID: {mal_id} (from anime_list.json)")
                        url = f"https://api.jikan.moe/v4/anime/{mal_id}"
                    else:
                        print(f"      No direct MAL ID match for '{anime_name}'. Falling back to search...")
                        # Priority search for exact title without members sort if it might lead to irrelevant popular shows
                        url = f"https://api.jikan.moe/v4/anime?q={anime_name}&limit=1"
                        
                    resp = requests.get(url, timeout=10)
                    resp.raise_for_status()
                    data = resp.json()
                    
                    img_url = None
                    if mal_id and data.get("data"):
                        img_url = data["data"]["images"]["jpg"]["large_image_url"]
                    elif data.get("data") and len(data["data"]) > 0:
                        img_url = data["data"][0]["images"]["jpg"]["large_image_url"]
                        
                    if img_url:
                        img_resp = requests.get(img_url, timeout=15)
                        img_resp.raise_for_status()
                        
                        with open(output_file_path, 'wb') as f:
                            f.write(img_resp.content)
                        time.sleep(1) # Rate limit exactly as requested
                    else:
                        print(f"      No results found in Jikan for {anime_name}.")
                except Exception as e:
                    print(f"      Fetch failed for {anime_name} at {url}: {e}")

    print("  Fetching missing Character Names from Jikan API by CHAR_ID...")
    # Jména postav se nemění — jednou stažené jméno drží lokální cache
    # (tools/top_chars_cache.json) a API se volá jen pro nová CHAR_ID.
    # Dřív se všech 10 jmen tahalo při každém běhu se změněným sešitem.
    char_cache_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "top_chars_cache.json")
    try:
        with open(char_cache_path, 'r', encoding='utf-8') as f:
            char_cache = json.load(f)
    except Exception:
        char_cache = {}
    char_cache_zmenena = False
    jikan_selhani_po_sobe = 0

    for sdata in top10_chars:
        char_id = sdata["data"].get("CHAR_ID")
        # Shape text je často jen název anime, skutečné jméno má MAL podle CHAR_ID
        if not char_id:
            continue

        cached_name = char_cache.get(char_id)
        if cached_name:
            sdata["data"]["ANIME_NAME"] = sdata["data"].get("NAME", "")
            sdata["data"]["NAME"] = cached_name
            continue

        if jikan_health.je_vypadek():
            print(f"    Jikan je mimo provoz (sdílená pojistka), CHAR_ID {char_id} se zkusí příště.")
            sdata["data"]["NAME"] = "Unknown Character"
            continue

        print(f"    Fetching Jikan Character API for ID: {char_id}")
        try:
            url = f"https://api.jikan.moe/v4/characters/{char_id}"
            resp = requests.get(url, timeout=10)
            resp.raise_for_status()
            data = resp.json()

            if data.get("data") and data["data"].get("name"):
                sdata["data"]["ANIME_NAME"] = sdata["data"].get("NAME", "")
                sdata["data"]["NAME"] = data["data"]["name"]
                char_cache[char_id] = data["data"]["name"]
                char_cache_zmenena = True
                jikan_selhani_po_sobe = 0
                time.sleep(1)
        except Exception as e:
            print(f"      Character Fetch failed for {char_id}: {e}")
            sdata["data"]["NAME"] = "Unknown Character"
            jikan_selhani_po_sobe += 1
            if jikan_selhani_po_sobe >= 3:
                jikan_health.nahlas_vypadek()
                print("      3 selhání po sobě: hlásím výpadek Jikanu do sdílené pojistky.")

    if char_cache_zmenena:
        try:
            with open(char_cache_path, 'w', encoding='utf-8') as f:
                json.dump(char_cache, f, ensure_ascii=False, indent=1)
        except Exception as e:
            print(f"  Warning: nelze uložit cache jmen postav: {e}")

    result = {
        "top10_anime": top10_anime,
        "hm_anime": hm_anime,
        "top10_chars": top10_chars
    }
    
    # Save hash to indicate this version has been extracted
    if current_hash:
        try:
            with open(hash_path, 'w') as f:
                f.write(current_hash)
        except:
            pass

    return result

def run_krok(popis, argv, cwd=None):
    """Spustí podskript a vypíše, jak dlouho běžel — regrese jsou pak vidět hned."""
    t = time.monotonic()
    subprocess.run(argv, check=True, cwd=cwd)
    print(f"[TIMER] {popis}: {time.monotonic() - t:.1f}s")


def main():
    t_total = time.monotonic()
    # --no-push = testovací běh: všechno se vygeneruje, ale bez commitu a push
    no_push = "--no-push" in sys.argv[1:]
    pos_args = [a for a in sys.argv[1:] if not a.startswith("--")]

    if len(pos_args) > 0:
        file_path = pos_args[0]
    else:
        # Cesta se odvozuje od umístění skriptu, ne napevno — celá složka
        # Anime_List se tak dá přesunout nebo přejmenovat a export běží dál.
        # Přebít jde proměnnou prostředí ANIME_LIST_ROOT nebo argumentem.
        file_path = os.path.join(anime_list_root(), "Anime list.xlsm")


    if len(pos_args) > 1:
        output_dir = pos_args[1]
    else:
        # Skript žije v anime-list-web/tools/ → data jsou o úroveň výš (anime-list-web/public/data)
        output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "public", "data")

    print(f"Loading Excel file: {file_path}")
    
    # Excel locks the active file (PermissionError). We must copy it to a temp file to read it.
    temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsm")
    os.close(temp_fd) # Close the file descriptor so we can overwrite it
    
    try:
        shutil.copy2(file_path, temp_path)
        # We don't use read_only=True anymore because it strips hyperlinks resulting in AttributeError
        # But we're safe because we copy the file first
        #
        # Obě podoby sešitu (hodnoty + vzorce/komentáře) se načítají souběžně
        # ve dvou vláknech. Zip dekomprese GIL pouští, XML parsování ne, takže
        # zisk není poloviční, ale je zadarmo: stejná knihovna, stejná data.
        t_load = time.monotonic()
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=2) as _pool:
            _f_data = _pool.submit(openpyxl.load_workbook, temp_path, data_only=True)
            _f_full = _pool.submit(openpyxl.load_workbook, temp_path, data_only=False)
            wb = _f_data.result()
            wb_comments = _f_full.result()
        print(f"[TIMER] Načtení obou sešitů (souběžně): {time.monotonic() - t_load:.1f}s")
        
        os.makedirs(output_dir, exist_ok=True)
        
        # Export each dataset
        print("Exporting Anime List (with comments)...")
        anime_list = export_anime_list(wb, wb_comments)
        with open(os.path.join(output_dir, "anime_list.json"), "w", encoding="utf-8") as f:
            json.dump(anime_list, f, ensure_ascii=False, indent=2)
        print(f"  Exported {len(anime_list)} anime entries")
        
        print("Exporting History Log...")
        # Load workbook again without data_only to get comments for history rewatches
        history = export_history_log(wb, wb_comments)
        with open(os.path.join(output_dir, "history_log.json"), "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        print(f"  Exported {len(history)} history entries")
        
        print("Exporting General Stats (with comments)...")
        stats = export_general_stats(wb, wb_comments)
        wb_comments.close() # Now safe to close after both anime_list and stats are done
        with open(os.path.join(output_dir, "stats.json"), "w", encoding="utf-8") as f:
            json.dump(stats, f, ensure_ascii=False, indent=2)
        
        print("Exporting Favorites...")
        favorites = export_favorites(wb)
        with open(os.path.join(output_dir, "favorites.json"), "w", encoding="utf-8") as f:
            json.dump(favorites, f, ensure_ascii=False, indent=2)
        print(f"  Exported {len(favorites)} favorite entries")
        
        print("Exporting Favorite OST Tables...")
        favorites_ost = export_ost_tables(wb, wb_comments)
        with open(os.path.join(output_dir, "favorites_ost.json"), "w", encoding="utf-8") as f:
            json.dump(favorites_ost, f, ensure_ascii=False, indent=2)
        print(f"  Exported OST Scenes: {len(favorites_ost['scenes'])}, Pieces: {len(favorites_ost['pieces'])}, Whole: {len(favorites_ost['whole'])}")
        
        print("Exporting Plan to Watch...")
        ptw = export_plan_to_watch(wb)
        with open(os.path.join(output_dir, "plan_to_watch.json"), "w", encoding="utf-8") as f:
            json.dump(ptw, f, ensure_ascii=False, indent=2)
        print(f"  Exported {len(ptw)} plan to watch entries")
        
        print("Exporting Category Ratings...")
        cat_ratings = export_category_ratings(wb)
        with open(os.path.join(output_dir, "category_ratings.json"), "w", encoding="utf-8") as f:
            json.dump(cat_ratings, f, ensure_ascii=False, indent=2)
        print(f"  Exported {len(cat_ratings)} anime with category ratings")
        
        print("Exporting Episode Ratings...")
        ep_ratings = export_episode_ratings(wb)
        with open(os.path.join(output_dir, "episode_ratings.json"), "w", encoding="utf-8") as f:
            json.dump(ep_ratings, f, ensure_ascii=False, indent=2)
        print(f"  Exported {len(ep_ratings)} anime with episode ratings")
        
        print("Exporting Notes/Reviews...")
        notes = export_notes(wb)
        with open(os.path.join(output_dir, "notes.json"), "w", encoding="utf-8") as f:
            json.dump(notes, f, ensure_ascii=False, indent=2)
        print(f"  Exported {len(notes)} anime with notes/reviews")
        
        print("Exporting Top Favorites & Characters...")
        top_favorites = export_top_favorites(file_path, output_dir) # Use original file_path for COM, wb is internal memory
        json_tf_path = os.path.join(output_dir, "top_favorites.json")
        has_items = any(len(top_favorites.get(k, [])) > 0 for k in ["top10_anime", "hm_anime", "top10_chars"])
        if has_items or not os.path.exists(json_tf_path):
            with open(json_tf_path, "w", encoding="utf-8") as f:
                json.dump(top_favorites, f, ensure_ascii=False, indent=2)
            print(f"  Exported Top 10 Anime: {len(top_favorites['top10_anime'])}, HM Anime: {len(top_favorites['hm_anime'])}, Top 10 Chars: {len(top_favorites['top10_chars'])}")
        else:
            print("  Warning: Top Favorites extraction returned 0 items, preserving existing top_favorites.json cache.")
    
        # Export metadata for version checking
        print("Exporting Metadata...")
        metadata = {
            "lastUpdated": int(time.time() * 1000)
        }
        with open(os.path.join(output_dir, "metadata.json"), "w", encoding="utf-8") as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)
        
        print("\nDone! Data exported to:", output_dir)
        
        script_root = os.path.dirname(os.path.abspath(__file__))

        # 1. Map thumbnails from folder
        print("Running map_from_folder.py to restore thumbnail paths...")
        try:
            script_path = os.path.join(script_root, "map_from_folder.py")
            run_krok("map_from_folder.py", [sys.executable, script_path], cwd=script_root)
            print("Thumbnails mapped successfully!")
        except Exception as e:
            print(f"Failed to run map_from_folder.py: {e}")

        # 2. Zbytek podskriptů běží souběžně v „lajnách" (B2). Pravidla:
        #  - oba Jikan skripty sdílí rate limit 3/s a 60/min, proto jedou za
        #    sebou v jedné lajně (postery navíc čtou čerstvou jikan_cache),
        #  - build_cover_thumbs čte Spotify obaly, běží až po extract_spotify_images,
        #  - výstup lajny se bufferuje a vypíše vcelku, ať se logy nemíchají,
        #  - git commit přijde až po doběhnutí všech lajn.
        def _spust_lajnu(kroky):
            """Spustí kroky lajny sériově; vrátí nasbíraný výstup vcelku."""
            buf = []
            env_potomka = os.environ.copy()
            env_potomka["PYTHONIOENCODING"] = "utf-8"
            for popis, argv in kroky:
                t = time.monotonic()
                try:
                    res = subprocess.run(argv, cwd=script_root, check=True,
                                         capture_output=True, text=True,
                                         encoding="utf-8", errors="replace",
                                         env=env_potomka)
                    vystup = (res.stdout or "") + (res.stderr or "")
                    buf.append(f"--- {popis} ---\n{vystup}"
                               f"[TIMER] {popis}: {time.monotonic() - t:.1f}s\n")
                except subprocess.CalledProcessError as e:
                    vystup = (e.stdout or "") + (e.stderr or "")
                    buf.append(f"--- {popis} SELHAL (kód {e.returncode}) ---\n{vystup}\n")
                except Exception as e:
                    buf.append(f"--- {popis} SELHAL ---\n{e}\n")
            return "".join(buf)

        # IMDb refresh se rozhoduje předem (týdenní odstup, viz komentář níž),
        # a když je na řadě, přidá se jako vlastní lajna.
        # Skript stahuje z IMDb datové sady o desítkách MB, takže nemá smysl ho
        # pouštět při každém exportu. Týdenní odstup stačí: epizodní známky se
        # mění pomalu a u právě vysílaných sérií se dopočítají příště.
        imdb_kroky = []
        try:
            imdb_cache_file = os.path.join(output_dir, "imdb_cache.json")
            stari_dnu = None
            if os.path.exists(imdb_cache_file):
                stari_dnu = (time.time() - os.path.getmtime(imdb_cache_file)) / 86400
            if stari_dnu is not None and stari_dnu < 7:
                print(f"IMDb cache je stara {stari_dnu:.1f} dne, refresh se preskakuje "
                      f"(obnovuje se jednou za 7 dni).")
            else:
                duvod = "cache neexistuje" if stari_dnu is None else f"cache je stara {stari_dnu:.1f} dne"
                print(f"IMDb cache: spusti se download_imdb_cache.py ({duvod}).")
                imdb_kroky.append(("download_imdb_cache.py",
                                   [sys.executable, os.path.join(script_root, "download_imdb_cache.py")]))
        except Exception as e:
            # Selhani kontroly nesmi shodit export, web pojede na starsi cachi.
            print(f"  Varovani: kontrola stari IMDb cache selhala: {e}")

        def _krok(nazev):
            return (nazev, [sys.executable, os.path.join(script_root, nazev)])

        lajny = [
            [_krok("download_jikan_cache.py"), _krok("download_journey_posters.py")],
            [_krok("extract_spotify_images.py"), _krok("build_cover_thumbs.py")],
            [_krok("export_docx_categories.py")],
            [_krok("build_ytmusic_ost.py")],
            [_krok("download_animethemes_cache.py")],
            [_krok("build_top_favorites_thumbs.py")],
        ]
        if imdb_kroky:
            lajny.append(imdb_kroky)

        print("Spouštím podskripty souběžně (výpisy přijdou po blocích)...")
        t_lajny = time.monotonic()
        from concurrent.futures import ThreadPoolExecutor, as_completed
        with ThreadPoolExecutor(max_workers=4) as pool:
            futures = [pool.submit(_spust_lajnu, kroky) for kroky in lajny]
            for f in as_completed(futures):
                print(f.result(), end="", flush=True)
        print(f"[TIMER] Vsechny podskripty (soubezne): {time.monotonic() - t_lajny:.1f}s")

        # 9. Push to GitHub
        # `-c gc.auto=0` je tu schvalne. Repozitar lezi ve slozce OneDrive,
        # ktera na souborech drzi zamky, takze automaticky uklid gitu neuspeje
        # smazat uz zabalene volne objekty a zeptá se
        # „Deletion of directory '.git/objects/xx' failed. Should I try again? (y/n)".
        # Export bezi bez konzole, na kterou by slo odpovedet, takze cely
        # skript na tom uvizne. Uklid tedy pri automatickem behu vypiname,
        # rucni `git gc` funguje dal beze zmeny.
        GIT = ["git", "-c", "gc.auto=0"]
        if no_push:
            print("Přepínač --no-push: commit a push se přeskakují (testovací běh).")
        else:
            print("Pushing data to GitHub...")
            try:
                web_dir = os.path.abspath(os.path.join(output_dir, "..", ".."))
                subprocess.run(GIT + ["add", "public/data/*", "public/images/*"], cwd=web_dir, check=True)
                # Když se reálně nic nezměnilo, nemá smysl commitovat a spouštět
                # nasazení. Samotný timestamp v metadata.json se za změnu
                # nepočítá a vrátí se zpět, ať klientům zbytečně neinvaliduje cache.
                res = subprocess.run(GIT + ["diff", "--cached", "--name-only"],
                                     cwd=web_dir, check=True, capture_output=True, text=True)
                zmenene = [l.strip() for l in res.stdout.splitlines() if l.strip()]
                if not zmenene or zmenene == ["public/data/metadata.json"]:
                    subprocess.run(GIT + ["reset", "-q", "--", "public/data", "public/images"],
                                   cwd=web_dir, check=False)
                    if zmenene:
                        subprocess.run(GIT + ["checkout", "-q", "--", "public/data/metadata.json"],
                                       cwd=web_dir, check=False)
                    print("Data se nezměnila, commit a nasazení se přeskakují.")
                else:
                    subprocess.run(GIT + ["commit", "-m", "Auto-update dat z Excelu (Background)"], cwd=web_dir, check=True)
                    subprocess.run(GIT + ["push", "origin", "main"], cwd=web_dir, check=True)
                    print("Git push completed successfully!")
            except Exception as e:
                try:
                    # Záchranná větev: schválně už jen public/, žádné `git add -A`.
                    # Dřívější add -A umělo vzít i rozpracované zdrojáky a rovnou
                    # je nasadit na web.
                    subprocess.run(GIT + ["add", "public/data", "public/images"], cwd=web_dir, check=True)
                    subprocess.run(GIT + ["commit", "-m", "Auto-update dat z Excelu (Background Fallback)"], cwd=web_dir, check=True)
                    subprocess.run(GIT + ["push", "origin", "main"], cwd=web_dir, check=True)
                    print("Git fallback push completed successfully!")
                except Exception as ge:
                    print(f"Failed to push to GitHub: {ge}")

    finally:
        print(f"[TIMER] Celkem export_data.py: {time.monotonic() - t_total:.1f}s")
        # Cleanup the temp file
        if 'wb' in locals():
            wb.close()
        try:
            os.remove(temp_path)
        except Exception as e:
            print(f"Warning: Could not remove temp file {temp_path}: {e}")

if __name__ == "__main__":
    main()

